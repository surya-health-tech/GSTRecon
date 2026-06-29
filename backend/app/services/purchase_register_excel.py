"""Parse purchase register Excel uploads (.xls / .xlsx)."""

from __future__ import annotations

import io
from pathlib import Path

from fastapi import HTTPException, status
from openpyxl import load_workbook

SUPPORTED_EXTENSIONS = frozenset({".xls", ".xlsx"})


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_xlsx(content: bytes, sheet_name: str | None) -> tuple[list[str], list[str], dict[str, str]]:
    try:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not read Excel file. Ensure it is a valid .xlsx workbook.",
        ) from exc

    sheets = wb.sheetnames
    if not sheets:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel file has no worksheets.")

    selected = sheet_name if sheet_name and sheet_name in sheets else sheets[0]
    ws = wb[selected]

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Selected sheet is empty or has no readable column headers.",
        )

    columns: list[str] = []
    for cell in header_row:
        label = _cell_to_str(cell)
        if label:
            columns.append(label)

    if not columns:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No column headers were found in the first row of the selected sheet.",
        )

    sample_row: dict[str, str] = {col: "" for col in columns}
    data_row = next(rows_iter, None)
    if data_row:
        for idx, col in enumerate(columns):
            if idx < len(data_row):
                sample_row[col] = _cell_to_str(data_row[idx])

    wb.close()
    return sheets, columns, sample_row


def _parse_xls(content: bytes, sheet_name: str | None) -> tuple[list[str], list[str], dict[str, str]]:
    try:
        import xlrd
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Legacy .xls support is not available on the server.",
        ) from exc

    try:
        book = xlrd.open_workbook(file_contents=content)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not read Excel file. Ensure it is a valid .xls workbook.",
        ) from exc

    sheets = book.sheet_names()
    if not sheets:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel file has no worksheets.")

    selected = sheet_name if sheet_name and sheet_name in sheets else sheets[0]
    sheet = book.sheet_by_name(selected)

    if sheet.nrows < 1:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Selected sheet is empty or has no readable column headers.",
        )

    columns: list[str] = []
    for col_idx in range(sheet.ncols):
        label = _cell_to_str(sheet.cell_value(0, col_idx))
        if label:
            columns.append(label)

    if not columns:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No column headers were found in the first row of the selected sheet.",
        )

    sample_row: dict[str, str] = {col: "" for col in columns}
    if sheet.nrows > 1:
        for idx, col in enumerate(columns):
            if idx < sheet.ncols:
                sample_row[col] = _cell_to_str(sheet.cell_value(1, idx))

    return sheets, columns, sample_row


def parse_excel_workbook(
    content: bytes,
    filename: str,
    sheet_name: str | None = None,
) -> tuple[list[str], str, list[str], dict[str, str]]:
    ext = Path(filename).suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Upload a .xls or .xlsx file.",
        )

    if ext == ".xlsx":
        sheets, columns, sample_row = _parse_xlsx(content, sheet_name)
    else:
        sheets, columns, sample_row = _parse_xls(content, sheet_name)

    selected = sheet_name if sheet_name and sheet_name in sheets else sheets[0]
    return sheets, selected, columns, sample_row


def parse_sheet_all_rows(
    content: bytes,
    filename: str,
    sheet_name: str | None = None,
) -> tuple[list[str], list[dict[str, str]]]:
    """Read all data rows from a purchase register worksheet (row 1 headers)."""
    ext = Path(filename).suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Upload a .xls or .xlsx file.",
        )

    if ext == ".xlsx":
        try:
            wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Could not read Excel file. Ensure it is a valid .xlsx workbook.",
            ) from exc
        sheets = wb.sheetnames
        if not sheets:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel file has no worksheets.")
        selected = sheet_name if sheet_name and sheet_name in sheets else sheets[0]
        ws = wb[selected]
        rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
        wb.close()
    else:
        try:
            import xlrd
        except ImportError as exc:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Legacy .xls support is not available on the server.",
            ) from exc
        try:
            book = xlrd.open_workbook(file_contents=content)
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Could not read Excel file. Ensure it is a valid .xls workbook.",
            ) from exc
        sheets = book.sheet_names()
        if not sheets:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel file has no worksheets.")
        selected = sheet_name if sheet_name and sheet_name in sheets else sheets[0]
        sheet = book.sheet_by_name(selected)
        rows = [
            tuple(sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols))
            for row_idx in range(sheet.nrows)
        ]

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Selected sheet is empty or has no readable column headers.",
        )

    header_row = rows[0]
    columns: list[str] = []
    for cell in header_row:
        label = _cell_to_str(cell)
        if label:
            columns.append(label)

    if not columns:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No column headers were found in the first row of the selected sheet.",
        )

    data_rows: list[dict[str, str]] = []
    for row in rows[1:]:
        if not any(_cell_to_str(cell) for cell in row):
            continue
        record = {col: "" for col in columns}
        for idx, col in enumerate(columns):
            if idx < len(row):
                record[col] = _cell_to_str(row[idx])
        data_rows.append(record)
    return columns, data_rows
