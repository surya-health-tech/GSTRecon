"""Export reconciliation case tabs to Excel."""

from __future__ import annotations

import io
import re
from typing import Any

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models import Client, ReconciliationCaseRecord
from app.services.reconciliation_cases import get_case

EXPORT_TABS = frozenset(
    {
        "summary",
        "all",
        "matched_with_books",
        "amount_mismatch",
        "portal_open",
        "books_open",
    }
)

TAB_LABELS: dict[str, str] = {
    "summary": "Summary",
    "all": "All Records",
    "matched_with_books": "Matched with Books",
    "amount_mismatch": "Amount Mismatch",
    "portal_open": "Portal Open Items",
    "books_open": "Books Open Items",
}

CATEGORY_LABELS: dict[str, str] = {
    "matched_with_books": "Matched with Books",
    "amount_mismatch": "Amount Mismatch",
    "portal_open": "Portal Open Items",
    "books_open": "Books Open Items",
}

MASTER_FIELDS: list[tuple[str, str, bool]] = [
    ("Voucher Number", "voucher_number", False),
    ("Company GSTIN", "company_gstin", False),
    ("Supplier Name", "supplier_name", False),
    ("Supplier GSTIN", "supplier_gstin", False),
    ("Supplier Invoice No", "supplier_invoice_no", False),
    ("Supplier Invoice Date", "supplier_invoice_date", False),
    ("Taxable Amount", "taxable_amount", True),
    ("IGST", "igst", True),
    ("SGST", "sgst", True),
    ("CGST", "cgst", True),
    ("Total Tax", "total_tax", True),
    ("Grand Total", "grand_total", True),
]

MONTH_NAMES = [
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
]


def validate_export_tab(tab: str) -> str:
    cleaned = tab.strip().lower()
    if cleaned not in EXPORT_TABS:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid export tab")
    return cleaned


def _safe_filename(case_name: str, tab: str) -> str:
    base = re.sub(r"[^\w.\- ]", "_", case_name).strip() or "case"
    tab_part = re.sub(r"[^\w.\- ]", "_", TAB_LABELS.get(tab, tab)).strip()
    return f"{base[:60]}_{tab_part}.xlsx"


def _field_value(data: dict[str, Any] | None, key: str, numeric: bool) -> str | float | None:
    if not data:
        return None
    if key == "total_tax" and numeric:
        total = data.get("total_tax")
        if total not in (None, "", 0, 0.0):
            try:
                return float(total)
            except (TypeError, ValueError):
                pass
        try:
            return float(data.get("igst") or 0) + float(data.get("cgst") or 0) + float(data.get("sgst") or 0)
        except (TypeError, ValueError):
            return 0.0
    value = data.get(key)
    if numeric:
        try:
            return float(value or 0)
        except (TypeError, ValueError):
            return 0.0
    text = str(value or "").strip()
    return text or None


def _autosize_columns(ws, max_width: int = 48) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            length = max(length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(length + 2, 12), max_width)


def _write_summary_sheet(ws, case, client_name: str | None) -> None:
    month_label = MONTH_NAMES[case.tax_period_month - 1] if 1 <= case.tax_period_month <= 12 else str(case.tax_period_month)
    tax_period = f"{month_label} {case.tax_period_year} ({case.tax_period_month:02d}/{case.tax_period_year})"
    summary = case.summary_counts or {}

    rows: list[list[Any]] = [
        ["Case Name", case.case_name],
        ["Client", client_name or ""],
        ["Tax Period", tax_period],
        ["Status", case.status.replace("_", " ").title()],
        ["GSTR-2B File", case.gstr2b_original_filename or ""],
        ["Purchase Register File", case.pr_original_filename or ""],
        ["GSTR-2B Mapping", case.gstr2b_mapping_name or ""],
        ["Purchase Register Mapping", case.pr_mapping_name or ""],
        [],
        ["Summary Metric", "Count"],
        ["Total GSTR-2B Records", summary.get("total_gstr2b_records", 0)],
        ["Total Purchase Register Records", summary.get("total_purchase_register_records", 0)],
        ["Matched with Books", summary.get("matched_with_books", 0)],
        ["Amount Mismatch", summary.get("amount_mismatch", 0)],
        ["Portal Open Items", summary.get("portal_open_items", 0)],
        ["Books Open Items", summary.get("books_open_items", 0)],
    ]

    header_font = Font(bold=True)
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 1 and value and value != "Summary Metric":
                cell.font = header_font
        if row and row[0] == "Summary Metric":
            for col_idx in (1, 2):
                ws.cell(row=row_idx, column=col_idx).font = header_font

    _autosize_columns(ws)


def _record_headers() -> list[str]:
    headers = ["Category", "Match Status", "Remarks", "Source"]
    for label, _key, _numeric in MASTER_FIELDS:
        headers.append(f"GSTR-2B {label}")
        headers.append(f"Purchase Register {label}")
    return headers


def _record_row(record: ReconciliationCaseRecord) -> list[Any]:
    normalized = record.normalized or {}
    portal = record.portal_data
    book = record.book_data
    row: list[Any] = [
        CATEGORY_LABELS.get(record.category, record.category),
        record.match_status or "",
        record.remarks or "",
        normalized.get("source") or "",
    ]
    for _label, key, numeric in MASTER_FIELDS:
        row.append(_field_value(portal, key, numeric))
        row.append(_field_value(book, key, numeric))
    return row


def _write_records_sheet(ws, records: list[ReconciliationCaseRecord]) -> None:
    headers = _record_headers()
    header_font = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font

    for row_idx, record in enumerate(records, start=2):
        for col_idx, value in enumerate(_record_row(record), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.freeze_panes = "A2"
    _autosize_columns(ws)


def export_case_tab_excel(
    db: Session,
    tenant_id: int,
    case_id: int,
    tab: str,
    *,
    client_name: str | None = None,
) -> tuple[bytes, str]:
    validated_tab = validate_export_tab(tab)
    case = get_case(db, tenant_id, case_id)

    if client_name is None and case.client_id:
        client = db.query(Client.client_name).filter(Client.id == case.client_id).first()
        client_name = client.client_name if client else None

    wb = Workbook()
    ws = wb.active
    ws.title = TAB_LABELS[validated_tab][:31]

    if validated_tab == "summary":
        _write_summary_sheet(ws, case, client_name)
    else:
        q = db.query(ReconciliationCaseRecord).filter(ReconciliationCaseRecord.case_id == case.id)
        if validated_tab != "all":
            q = q.filter(ReconciliationCaseRecord.category == validated_tab)
        records = q.order_by(ReconciliationCaseRecord.id.asc()).all()
        _write_records_sheet(ws, records)

    buffer = io.BytesIO()
    wb.save(buffer)
    filename = _safe_filename(case.case_name, validated_tab)
    return buffer.getvalue(), filename
