"""Parse GSTR-2B Excel uploads and match workbook sheets to supported tabs."""

from __future__ import annotations

import io
import re
from pathlib import Path

from fastapi import HTTPException, status
from openpyxl import load_workbook

from app.services.purchase_register_excel import SUPPORTED_EXTENSIONS, _cell_to_str

SUPPORTED_GSTR2B_TABS: tuple[str, ...] = (
    "B2B",
    "B2BA",
    "B2B-CDNR",
    "B2B-CDNRA",
    "ECO",
    "ECOA",
    "ISD",
    "ISDA",
    "IMPG",
    "IMPGA",
    "IMPSEZ",
    "IMPSEZA",
)

CDNR_TABS = frozenset({"B2B-CDNR", "B2B-CDNRA"})

# GST portal workbooks sometimes label IMPSEZ sheets as IMPGSEZ.
_TAB_ALIASES: dict[str, str] = {
    "impgsez": "IMPSEZ",
    "impsez": "IMPSEZ",
    "impgseza": "IMPSEZA",
    "impseza": "IMPSEZA",
}

_HEADER_KEYWORDS = (
    "gstin",
    "invoice",
    "trade/legal",
    "legal name",
    "note number",
    "note type",
    "document number",
    "document date",
    "taxable",
    "integrated tax",
    "central tax",
    "state/ut tax",
    "icegate",
    "bill of entry",
    "port code",
    "isd document",
    "supply attract",
    "place of supply",
    "original details",
    "revised details",
)

_GSTIN_RE = re.compile(r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][0-9A-Z]Z[0-9A-Z]$")
_DATE_RE = re.compile(r"^\d{1,2}/\d{1,2}/\d{4}$")


def _normalize_tab_key(name: str) -> str:
    cleaned = name.strip()
    if cleaned.lower().endswith(" extract"):
        cleaned = cleaned[: -len(" extract")].strip()
    return re.sub(r"[^a-z0-9]", "", cleaned.lower())


_TAB_LOOKUP: dict[str, str] = {_normalize_tab_key(tab): tab for tab in SUPPORTED_GSTR2B_TABS}
for alias_key, tab in _TAB_ALIASES.items():
    _TAB_LOOKUP[alias_key] = tab


def match_excel_sheet_to_tab(sheet_name: str) -> str | None:
    key = _normalize_tab_key(sheet_name)
    if key in _TAB_LOOKUP:
        return _TAB_LOOKUP[key]
    # Ignore reversal/rejected variants such as "B2B (ITC Reversal)".
    if "(" in sheet_name or "reversal" in sheet_name.lower() or "rejected" in sheet_name.lower():
        return None
    return None


def _looks_like_data_cell(value: str) -> bool:
    if not value:
        return False
    if _GSTIN_RE.match(value.upper()):
        return True
    if _DATE_RE.match(value):
        return True
    try:
        float(value.replace(",", ""))
        return True
    except ValueError:
        pass
    return False


def _row_text_cells(row: tuple[object, ...]) -> list[tuple[int, str]]:
    return [(idx, _cell_to_str(cell)) for idx, cell in enumerate(row) if _cell_to_str(cell)]


def _score_header_row(cells: list[tuple[int, str]]) -> int:
    if len(cells) < 2:
        return 0
    score = 0
    data_like = 0
    for _, text in cells:
        lower = text.lower()
        if _looks_like_data_cell(text):
            data_like += 1
            continue
        for kw in _HEADER_KEYWORDS:
            if kw in lower:
                score += 3
        if any(ch.isalpha() for ch in text):
            score += 1
    score -= data_like * 4
    return score


def _detect_header_block(rows: list[tuple[object, ...]]) -> tuple[list[int], int]:
    """Return header row indices (0-based) and first data row index."""
    scored: list[tuple[int, int]] = []
    for idx, row in enumerate(rows):
        cells = _row_text_cells(row)
        score = _score_header_row(cells)
        if score >= 6:
            scored.append((idx, score))

    if not scored:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not locate column headers in this GSTR-2B sheet.",
        )

    # Portal sheets use 1-2 consecutive header rows before data.
    header_indices = [scored[0][0]]
    for i in range(1, len(scored)):
        prev = scored[i - 1][0]
        cur = scored[i][0]
        if cur == prev + 1 and len(header_indices) < 3:
            header_indices.append(cur)
        elif cur > header_indices[-1] + 1:
            break

    data_start = header_indices[-1] + 1
    while data_start < len(rows):
        cells = _row_text_cells(rows[data_start])
        if not cells:
            data_start += 1
            continue
        # Skip repeated sub-header rows that occasionally appear without data.
        if _score_header_row(cells) >= 8 and not any(_looks_like_data_cell(t) for _, t in cells):
            header_indices.append(data_start)
            data_start += 1
            continue
        break

    if data_start >= len(rows):
        return header_indices, len(rows)

    return header_indices, data_start


def _parent_label_for_column(row: tuple[object, ...], col_idx: int) -> str | None:
    """Walk left to find merged parent header for a sub-column."""
    for idx in range(col_idx, -1, -1):
        label = _cell_to_str(row[idx]) if idx < len(row) else ""
        if label:
            return label
    return None


def _build_column_names(
    rows: list[tuple[object, ...]],
    header_indices: list[int],
) -> tuple[list[str], list[int]]:
    max_cols = max(len(rows[i]) for i in header_indices)
    names: list[str] = []
    indices: list[int] = []
    seen: dict[str, int] = {}

    for col_idx in range(max_cols):
        parts: list[str] = []
        for row_idx in header_indices:
            row = rows[row_idx]
            label = _cell_to_str(row[col_idx]) if col_idx < len(row) else ""
            if label:
                parts.append(label)

        if not parts:
            parent = _parent_label_for_column(rows[header_indices[0]], col_idx)
            if parent and parent.lower() not in {"tax amount", "invoice details", "document details"}:
                name = parent
            else:
                continue
        else:
            name = parts[-1]

        if name in seen:
            seen[name] += 1
            name = f"{name} ({seen[name]})"
        else:
            seen[name] = 1
        names.append(name)
        indices.append(col_idx)

    if not names:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No column headers were found in the detected header rows.",
        )
    return names, indices


def _parse_rows_to_columns_and_sample(
    rows: list[tuple[object, ...]],
) -> tuple[list[str], dict[str, str], int]:
    header_indices, data_start = _detect_header_block(rows)
    columns, col_indices = _build_column_names(rows, header_indices)
    sample_row: dict[str, str] = {col: "" for col in columns}

    if data_start < len(rows):
        data_row = rows[data_start]
        for name, col_idx in zip(columns, col_indices):
            if col_idx < len(data_row):
                sample_row[name] = _cell_to_str(data_row[col_idx])

    return columns, sample_row, header_indices[0] + 1


def _load_xlsx(content: bytes):
    try:
        return load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not read Excel file. Ensure it is a valid, unprotected .xlsx workbook.",
        ) from exc


def _load_xls(content: bytes):
    try:
        import xlrd
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Legacy .xls support is not available on the server.",
        ) from exc
    try:
        return xlrd.open_workbook(file_contents=content)
    except Exception as exc:
        msg = str(exc).lower()
        if "password" in msg or "encrypted" in msg:
            detail = "The Excel file appears to be password protected. Upload an unprotected file."
        else:
            detail = "Could not read Excel file. Ensure it is a valid .xls workbook."
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=detail) from exc


def list_workbook_sheets(content: bytes, filename: str) -> list[str]:
    ext = Path(filename).suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Upload a .xls or .xlsx file.",
        )
    if ext == ".xlsx":
        wb = _load_xlsx(content)
        sheets = list(wb.sheetnames)
        wb.close()
    else:
        book = _load_xls(content)
        sheets = list(book.sheet_names())
    if not sheets:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel file has no worksheets.")
    return sheets


def _read_sheet_rows_xlsx(wb, sheet_name: str, max_rows: int | None = 40) -> list[tuple[object, ...]]:
    ws = wb[sheet_name]
    if max_rows is None:
        return [tuple(row) for row in ws.iter_rows(values_only=True)]
    return [tuple(row) for row in ws.iter_rows(values_only=True, max_row=max_rows)]


def _read_sheet_rows_xls(book, sheet_name: str, max_rows: int | None = 40) -> list[tuple[object, ...]]:
    sheet = book.sheet_by_name(sheet_name)
    limit = sheet.nrows if max_rows is None else min(sheet.nrows, max_rows)
    rows: list[tuple[object, ...]] = []
    for row_idx in range(limit):
        rows.append(tuple(sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)))
    return rows


def parse_sheet_data(content: bytes, filename: str, sheet_name: str) -> tuple[list[str], dict[str, str]]:
    ext = Path(filename).suffix.lower()
    if ext == ".xlsx":
        wb = _load_xlsx(content)
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Sheet '{sheet_name}' was not found in the workbook.",
            )
        rows = _read_sheet_rows_xlsx(wb, sheet_name)
        wb.close()
    else:
        book = _load_xls(content)
        if sheet_name not in book.sheet_names():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Sheet '{sheet_name}' was not found in the workbook.",
            )
        rows = _read_sheet_rows_xls(book, sheet_name)

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Sheet '{sheet_name}' is empty or has no readable column headers.",
        )

    columns, sample_row, _ = _parse_rows_to_columns_and_sample(rows)
    return columns, sample_row


def parse_sheet_all_rows(
    content: bytes, filename: str, sheet_name: str
) -> tuple[list[str], list[dict[str, str]]]:
    """Read all data rows from a GSTR-2B worksheet using multi-row header detection."""
    ext = Path(filename).suffix.lower()
    if ext == ".xlsx":
        wb = _load_xlsx(content)
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Sheet '{sheet_name}' was not found in the workbook.",
            )
        rows = _read_sheet_rows_xlsx(wb, sheet_name, max_rows=None)
        wb.close()
    else:
        book = _load_xls(content)
        if sheet_name not in book.sheet_names():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Sheet '{sheet_name}' was not found in the workbook.",
            )
        rows = _read_sheet_rows_xls(book, sheet_name, max_rows=None)

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Sheet '{sheet_name}' is empty or has no readable column headers.",
        )

    header_indices, data_start = _detect_header_block(rows)
    columns, col_indices = _build_column_names(rows, header_indices)
    data_rows: list[dict[str, str]] = []
    for row_idx in range(data_start, len(rows)):
        row = rows[row_idx]
        if not any(_cell_to_str(cell) for cell in row):
            continue
        record = {name: "" for name in columns}
        for name, col_idx in zip(columns, col_indices):
            if col_idx < len(row):
                record[name] = _cell_to_str(row[col_idx])
        data_rows.append(record)
    return columns, data_rows


def map_workbook_sheets(excel_sheets: list[str]) -> dict[str, str | None]:
    """Map each supported tab to its matched Excel sheet name, if any."""
    result: dict[str, str | None] = {tab: None for tab in SUPPORTED_GSTR2B_TABS}
    used_excel: set[str] = set()
    for sheet in excel_sheets:
        tab = match_excel_sheet_to_tab(sheet)
        if tab and result[tab] is None and sheet not in used_excel:
            result[tab] = sheet
            used_excel.add(sheet)
    return result
